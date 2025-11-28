#!/usr/bin/env python3
"""
Intelligent CDAC Agent - Lets AI figure out the array structure through reasoning
No hard-coded rules, only basic observation tools
"""

import json
import sys
import os
from smolagents import Tool, CodeAgent, OpenAIServerModel
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula


class ObserveExcelTool(Tool):
    """Basic tool to observe Excel contents - no interpretation"""
    
    name = "observe_excel"
    description = """
    Observes and returns raw information about cells in an Excel file.
    Returns cell coordinates, values, and colors without any interpretation.
    Use this to understand what the user has created.
    
    Args:
        file_path: Path to Excel file or JSON
        region: Optional. Specify region like "A1:Z20" to focus on specific area
    
    Returns:
        Raw cell information as JSON
    """
    
    inputs = {
        "file_path": {"type": "string", "description": "Excel or JSON file path"},
        "sheet_name": {"type": "string", "description": "Sheet name (optional, defaults to first sheet)", "nullable": True},
        "region": {"type": "string", "description": "Optional region like 'A1:Z20'", "nullable": True}
    }
    output_type = "string"
    
    def _read_excel(self, file_path, sheet_name=None):
        """Read Excel file - reads TWICE to get both formulas and calculated values"""
        
        # First pass: Read formulas (data_only=False)
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        
        # Select sheet
        if sheet_name:
            if sheet_name in wb_formulas.sheetnames:
                ws_formulas = wb_formulas[sheet_name]
            else:
                available = ', '.join(wb_formulas.sheetnames)
                wb_formulas.close()
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")
        else:
            ws_formulas = wb_formulas[wb_formulas.sheetnames[0]]
        
        formulas_dict = {}
        for row in ws_formulas.iter_rows():
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    # Check if it's a formula
                    if isinstance(value, str) and value.startswith('='):
                        formulas_dict[cell.coordinate] = value
                    # Handle ArrayFormula objects
                    elif hasattr(value, '__class__') and 'ArrayFormula' in str(type(value)):
                        if hasattr(value, 'text') and value.text:
                            formulas_dict[cell.coordinate] = str(value.text)
                        elif hasattr(value, 'ref'):
                            formulas_dict[cell.coordinate] = f"={value.ref}"
        wb_formulas.close()
        
        # Second pass: Read calculated values (data_only=True)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Select same sheet
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
        
        cells_data = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None and not any([cell.fill, cell.font]):
                    continue
                
                coord = cell.coordinate
                cell_info = {'row': cell.row, 'col': cell.column}
                
                # Store the calculated value
                if cell.value is not None:
                    cell_info['value'] = str(cell.value)
                
                # Add formula if exists
                if coord in formulas_dict:
                    cell_info['formula'] = formulas_dict[coord]
                
                if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    fg = cell.fill.fgColor
                    try:
                        # Check RGB color (only if it's a valid string)
                        if hasattr(fg, 'rgb') and fg.rgb:
                            rgb_str = str(fg.rgb)
                            # Skip if it's an error message
                            if not rgb_str.startswith('Values must be'):
                                if len(rgb_str) == 8:  # ARGB
                                    cell_info['color_rgb'] = f"#{rgb_str[2:]}"
                                else:
                                    cell_info['color_rgb'] = f"#{rgb_str}" if not rgb_str.startswith('#') else rgb_str
                        
                        # Check theme color
                        if hasattr(fg, 'theme') and fg.theme is not None:
                            theme_val = fg.theme
                            # Convert to int
                            try:
                                cell_info['color_theme'] = int(theme_val)
                            except (TypeError, ValueError):
                                pass  # Skip invalid theme
                            
                            # Check tint
                            if hasattr(fg, 'tint') and fg.tint:
                                try:
                                    cell_info['color_tint'] = float(fg.tint)
                                except (TypeError, ValueError):
                                    pass
                    except Exception as e:
                        pass  # Skip color info if conversion fails
                
                cells_data[coord] = cell_info
        
        wb.close()
        return cells_data
    
    def forward(self, file_path: str, sheet_name: str = None, region: str = None) -> str:
        """Observe Excel contents"""
        # Read file
        if file_path.endswith('.json'):
            with open(file_path, 'r') as f:
                data = json.load(f)
                sheet = list(data['sheets'].values())[0]
                cells_data = sheet['cells']
        else:
            cells_data = self._read_excel(file_path, sheet_name=sheet_name)
        
        # Filter by region if specified
        if region:
            # Parse region like "A1:Z20"
            pass  # Simplified for now
        
        # Return summary
        return json.dumps({
            'total_cells': len(cells_data),
            'sample_cells': dict(list(cells_data.items())[:20]),
            'all_cells': cells_data
        }, indent=2)


# ============================================================================
# REMOVED TOOLS - AI does everything with Python code now!
# ============================================================================
# The following tools were removed because AI (CodeAgent) can do everything
# directly with Python code:
#
# - AnalyzeDataTool: AI analyzes data directly from observe_excel output
# - ExtractMetadataTool: AI extracts metadata with Python
# - CreateMappingTool: AI creates mappings directly in JSON output  
# - ValidateWithFormulasTool: AI validates using Python code
#
# Benefits:
# ✅ More flexible - AI adapts to any Excel format
# ✅ More intelligent - AI reasons about the data
# ✅ Simpler codebase - Only 1 tool instead of 4
# ============================================================================


def create_intelligent_agent():
    """Create agent with minimal tools - let AI reason and code freely"""
    
    # Only provide the essential tool - AI can do everything else with Python!
    tools = [
        ObserveExcelTool()  # Read Excel → AI analyzes with Python code
    ]
    
    model = OpenAIServerModel(
        model_id="deepseek-chat",
        api_base="https://api.deepseek.com/v1",
        api_key="sk-1b2d9b3e065b48d881680d9bc0b9d69e",
        flatten_messages_as_text=True,
        max_tokens=8192,
        temperature=0.3  # Lower temperature for more consistent reasoning
    )
    
    agent = CodeAgent(
        tools=tools,
        model=model,
        max_steps=20,  # More steps for reasoning
        additional_authorized_imports=[
            'os', 'pathlib', 'io', 'sys', 'subprocess', 
            'typing', 'posixpath', 'importlib', 'glob', 'json'
        ],
    )
    
    return agent


def main():
    if len(sys.argv) < 2:
        print("Usage: python intelligent_cdac_agent.py <excel_file> [sheet_name]")
        print("  excel_file: Path to Excel file")
        print("  sheet_name: (Optional) Specific sheet name to analyze")
        sys.exit(1)
    
    input_file = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    
    print("="*80)
    print("🤖 Intelligent CDAC Analysis Agent")
    print("="*80)
    print(f"\n📊 Input: {input_file}")
    if sheet_name:
        print(f"📑 Sheet: {sheet_name}")
    print(f"🧠 Model: DeepSeek Chat")
    print("\n💡 This agent will:")
    print("   1. Observe the Excel file")
    print("   2. Ask analytical questions to understand the structure")
    print("   3. Reason about which region is the CDAC array")
    print("   4. Understand how capacitors are grouped (by text, color, etc)")
    print("   5. Generate the correct mapping table")
    print("\n" + "="*80)
    
    agent = create_intelligent_agent()
    
    task = f"""
Analyze the Excel file '{input_file}' to understand the CDAC capacitor array and generate a JSON output for layout generation.

You are a CodeAgent - you can write Python code directly to analyze data!

**ULTIMATE GOAL**: 
Generate a structured JSON that can be passed to a layout generation agent.

🚨 **CRITICAL**: You MUST return the COMPLETE structured JSON data in a SINGLE response!
- Include ALL required fields: array_structure (with COMPLETE array_data), connection_groups (with positions), understanding
- 🚨 **MUST return ACTUAL JSON data** - NOT just text descriptions or summaries!
- The master agent needs the ACTUAL array_data (complete 2D array) and connection_groups (with positions) to generate layout
- DO NOT return only text descriptions like "Successfully analyzed..." - return the ACTUAL JSON structure!
- If you return text descriptions without actual JSON data, the master agent cannot process it and will call you again
- The complete JSON output must be ready for immediate use by the layout generation agent

**WORKFLOW**:

1. **START WITH FULL OBSERVATION** - Get ALL cells first (no need to probe):
   ```python
   # First call: Get ALL cells to understand the full structure
   data = observe_excel(file_path='{input_file}'{', sheet_name="' + sheet_name + '"' if sheet_name else ''})
   
   # This gives you the complete picture - all cells, formulas, colors
   parsed = json.loads(data)
   cells = parsed['all_cells']  # All cells are here!
   ```

2. **ANALYZE WITH PYTHON** - Use the full data to understand structure:
   ```python
   import json
   from collections import Counter, defaultdict
   
   # Analyze the full dataset:
   # - Find array boundaries (rows/columns)
   # - Group by value or color to identify connections
   # - Extract metadata (headers, labels, formulas)
   # - Understand the overall structure
   
   # Example: Find all unique values
   values = Counter([cell['value'] for cell in cells.values() if 'value' in cell])
   
   # 🚨 CRITICAL: AUTOMATICALLY DETECT array boundaries from ALL cells with 'value' field!
   # - Find ALL cells that have 'value' field (including "0" values - they are dummy capacitors!)
   # - Calculate min/max row and column numbers from these cells
   # - The array boundaries must include ALL rows and columns that have any value (including "0")
   ```

3. **FOCUSED ANALYSIS** (Optional - if needed):
   ```python
   # If you need to focus on a specific region for detailed analysis,
   # you can call observe_excel again with a region parameter:
   # data_region = observe_excel(file_path='...', region='C4:R22')
   # 
   # But usually the full data from step 1 is enough!
   ```
   
   **KEY POINT**: Start with full observation (no probing), then analyze. 
   You can call observe_excel multiple times if needed for focused analysis.

3. **UNDERSTAND SPECIAL MEANINGS** (During analysis):
   - **"0" = Dummy capacitors** (Dummy capacitors, not connected to actual circuit)
   - 🚨 **CRITICAL**: Cells with value "0" are STILL part of the array and must be included in array_data!
   - 🚨 **CRITICAL**: The array_data MUST include the complete dummy ring (all "0" values around the core array)!
   - The dummy ring typically forms the outer boundary of the array (leftmost column, rightmost column, top row, bottom row with "0" values)
   - **Fractional values like "2.3", "4.4"**: 
     * The NUMBER (2, 4) = actual count of connected capacitors
     * The VALUE (2.3, 4.4) = capacitance weight (redundancy/tuning)
     * Example: "2.3" means 2 capacitors connected, but weighted as 2.3 units
   - **Integer values like "128", "64"**: 
     * Direct binary-weighted capacitors
     * Number = both count AND weight

4. **VALIDATE WITH FORMULAS** (During analysis):
   Formula cells are VALIDATION DATA! Find them in `cells` and validate:
   ```python
   formula_cells = {{coord: cell for coord, cell in cells.items() if 'formula' in cell}}
   # COUNTIF → expected counts, SUM → expected totals
   # Validate your connection_groups match formula results!
   ```

5. **GENERATE STRUCTURED JSON OUTPUT** (CRITICAL):
   
   Output JSON with **REQUIRED** and **OPTIONAL** fields:
   
   ```json
   {{
     // ============ REQUIRED FIELDS ============
     
     "array_structure": {{
       "excel_region": {{"row_start": 4, "row_end": 22, "col_start": 3, "col_end": 18}},
       "dimensions": {{"rows": 19, "cols": 16}},
       "type": "binary_weighted",  // Array type: "binary_weighted", "unary_weighted", etc.
       "array_data": [
         ["0", "0", "0", "0", ...],  // Complete 2D array, MUST include dummy ring (all "0" values)
         ["0", "128", "128", "64", ...], 
         ["0", "128", "128", "64", ...],
         ...
       ]
       // 🚨 CRITICAL: array_data MUST include the complete dummy ring (all "0" values)!
     }},
     
     "understanding": "Detailed analysis of the CDAC array layout, connection relationships, special capacitor types, and design intent. This should be comprehensive and human-readable.",
     
     // ============ OPTIONAL FIELDS (based on actual situation) ============
     
     "connection_groups": [  // If connection relationships can be identified
       {{
         "value": "128",
         "weight": 128.0,
         "actual_count": 256,
         "positions": [[0,0], [0,1], ...],
         "description": "...",
         "pin_name": "MSB5"  // If Excel has corresponding pin information, add this field - IMPORTANT!!!
       }}
     ],
     
     "pin_assignments": {{  // If Excel has pin information
       "LSB": [...],
       "MSB": [...]
     }},
     
     "metadata": {{  // Any useful statistical information
       "total_positions": 304,
       "source_file": "{input_file}",
       ...
     }},
     
     "validation": {{  // If formula validation exists
       "method": "formula_cells",
       "accuracy": 100.0,
       "status": "VALIDATED",
       "details": "..."
     }}
   }}
   ```
   
   **CRITICAL REQUIREMENTS FOR JSON OUTPUT**:
   
   🚨 **MUST output STANDARD JSON format directly - no Python dict!**
   
   ✅ DO THIS:
   ```json
   {{
     "array_structure": {{
       "array_data": [[...]]
     }},
     "understanding": "..."
   }}
   ```
   
   ❌ NOT THIS (Python dict):
   ```python
   {{'array_structure': {{'array_data': [[...]]}}, 'understanding': '...'}}
   ```
   
   **Format Rules**:
   - ✅ Use double quotes `"` for all keys and string values
   - ✅ Use lowercase `true`, `false`, `null` (NOT Python's True/False/None)
   - ✅ Wrap output in ```json code block
   - ✅ **MUST HAVE**: `array_structure` + `understanding`
   - 🔧 **OPTIONAL**: `connection_groups`, `metadata`, `validation` (based on actual situation)
   - ✅ `array_data` is **pure 2D list** of strings
   - 🚨 **CRITICAL**: `positions` arrays MUST be COMPACT on ONE line:
     ✅ "positions": [[0,0], [0,1], [1,0], [1,1], [2,0], [2,1]]
     ❌ NOT: "positions": [[0,0],\n[0,1],\n[1,0],...] (too long!)
   - ✅ All positions use **0-based array indices**

**OUTPUT REQUIREMENTS**:
🚨 **CRITICAL**: Return COMPLETE structured data in ONE response - do NOT return partial data or summaries!

1. ✅ Valid JSON that can be parsed by downstream agent
2. ✅ **array_structure** with complete 2D array_data (REQUIRED)
   - 🚨 **CRITICAL**: array_data MUST include the complete dummy ring (all "0" values around the core array)!
   - MUST include the full 2D array with all rows and columns
3. ✅ **understanding** with comprehensive analysis (REQUIRED)
4. ✅ **connection_groups** (REQUIRED if connections can be identified)
   - MUST include all connection groups with their positions
   - If you can identify connections, you MUST include this field - do NOT omit it!
5. ✅ validation if formulas exist (OPTIONAL)
6. ✅ Other fields based on what's actually in the Excel (FLEXIBLE)

**REMEMBER**: The master agent expects COMPLETE data. If you return incomplete data (e.g., only understanding without array_data or connection_groups), the master agent will call you again, causing unnecessary repeated calls. Always return the FULL structured JSON in your first and only response!

**IMPORTANT INSIGHTS TO CAPTURE**:
- "0" capacitors are dummies - note their positions
- 🚨 **CRITICAL**: The dummy ring (all "0" values) MUST be included in array_data - they form the outer boundary of the array!
- Fractional values (2.3, 4.4, 8.4) indicate redundancy
  * Extract the integer part for actual capacitor count
  * The fractional value is the effective weight
- Binary-weighted main capacitors (128, 64, 32, 16, 8, 4, 2, 1)
- LSB/MSB pin assignments and their meanings
- **Formula cells = VALIDATION DATA** ⭐ 
  * COUNTIF → expected occurrence counts (MUST MATCH your analysis!)
  * SUM → expected totals
  * Every user's formulas may be in different locations - READ and UNDERSTAND them!
  * If validation fails, revise your analysis before outputting JSON

Remember: 
1. Write ANY Python code you need for analysis
2. **When building JSON output, ensure `positions` arrays are COMPACT**:
   ```python
   # When creating connection_groups, format positions compactly and add pin_name if available:
   group = {{
       "value": "128",
       "positions": [[0,0], [0,1], [1,0], [1,1]],  # ONE line, no line breaks!
       "pin_name": "MSB5"  # Add this field if Excel has matching pin assignment
   }}
   
   # Use json.dumps with separators to keep it compact:
   import json
   output = json.dumps(result, separators=(',', ':'), ensure_ascii=False)
   # This keeps positions on one line!
   ```
   ❌ DO NOT use indent=2 for positions arrays (makes file too long!)
   ✅ Keep positions arrays on ONE line: [[0,0], [0,1], [1,0], ...]
3. Output the final result as STANDARD JSON in a ```json code block
4. The JSON will be saved directly - no external formatting!
"""
    
    print("\n🎯 Task:")
    print(task)
    print("\n" + "="*80)
    print("🔄 Agent is analyzing...\n")
    
    try:
        result = agent.run(task)
        
        print("\n" + "="*80)
        print("✅ ANALYSIS COMPLETE")
        print("="*80)
        print("\n📋 Result:")
        print(result)
        
        # Save AI output directly - AI should output valid JSON
        output_json = os.path.splitext(input_file)[0] + '_cdac_layout.json'
        
        # Extract JSON from code block if present, otherwise use raw output
        result_str = str(result)
        if '```json' in result_str:
            json_start = result_str.find('```json') + 7
            json_end = result_str.find('```', json_start)
            json_content = result_str[json_start:json_end].strip()
        elif '```' in result_str:
            json_start = result_str.find('```') + 3
            json_end = result_str.find('```', json_start)
            json_content = result_str[json_start:json_end].strip()
        else:
            json_content = result_str
        
        # Save directly - AI is responsible for valid JSON format
        with open(output_json, 'w', encoding='utf-8') as f:
            f.write(json_content)
        
        print(f"\n💾 JSON saved to: {output_json}")
        print("   ✅ AI output saved directly - no external parsing!")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

